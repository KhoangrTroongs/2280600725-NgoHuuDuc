import pandas as pd
import random
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define categories
categories = ["Veston", "Quần tây", "Áo sơ mi"]

# Define product name templates for each category
veston_templates = [
    "Veston {color} {style} {material}",
    "Áo vest {color} {style} {fit}",
    "Bộ vest {color} {occasion} {material}",
    "Veston {brand} {color} {fit}",
    "Áo blazer {color} {style} {material}"
]

quan_tay_templates = [
    "Quần tây {color} {style} {material}",
    "Quần âu {color} {fit} {material}",
    "Quần tây {brand} {color} {fit}",
    "Quần âu {color} {style} {occasion}",
    "Quần tây {color} {material} {fit}"
]

ao_so_mi_templates = [
    "Áo sơ mi {color} {style} {material}",
    "Áo sơ mi {brand} {color} {fit}",
    "Áo sơ mi {color} {pattern} {occasion}",
    "Áo sơ mi {color} {sleeve} {material}",
    "Áo sơ mi {color} {style} {fit}"
]

# Define attributes for product names
colors = ["đen", "trắng", "xanh navy", "xanh dương", "xám", "be", "nâu", "xanh rêu", "đỏ đô", "xanh đen", "kẻ sọc", "họa tiết"]
styles = ["cổ điển", "hiện đại", "thanh lịch", "trẻ trung", "công sở", "dự tiệc", "casual", "slim fit", "regular fit"]
materials = ["len", "cotton", "linen", "polyester", "wool", "cashmere", "tweed", "kaki", "nhung", "denim"]
brands = ["Elegance", "Gentleman", "Classic", "Modern", "Premium", "Luxury", "Executive", "Business", "Formal"]
fits = ["ôm body", "regular fit", "slim fit", "oversize", "standard fit", "relaxed fit", "skinny fit"]
occasions = ["dự tiệc", "công sở", "cưới hỏi", "dạo phố", "lễ hội", "casual", "formal"]
patterns = ["trơn", "kẻ sọc", "kẻ caro", "họa tiết", "chấm bi", "in hoa", "thêu"]
sleeves = ["tay dài", "tay ngắn", "tay lỡ", "cộc tay"]

# Define price ranges for each category
price_ranges = {
    "Veston": (1500000, 5000000),
    "Quần tây": (500000, 1500000),
    "Áo sơ mi": (350000, 1200000)
}

# Define sizes
sizes = ["S", "M", "L", "XL"]

# Generate 100 random products
products = []

for i in range(100):
    # Randomly select a category
    category = random.choice(categories)
    
    # Select appropriate template based on category
    if category == "Veston":
        template = random.choice(veston_templates)
    elif category == "Quần tây":
        template = random.choice(quan_tay_templates)
    else:  # Áo sơ mi
        template = random.choice(ao_so_mi_templates)
    
    # Generate product name
    product_name = template.format(
        color=random.choice(colors),
        style=random.choice(styles),
        material=random.choice(materials),
        brand=random.choice(brands),
        fit=random.choice(fits),
        occasion=random.choice(occasions),
        pattern=random.choice(patterns),
        sleeve=random.choice(sleeves)
    )
    
    # Generate price within category range
    min_price, max_price = price_ranges[category]
    price = random.randint(min_price // 10000, max_price // 10000) * 10000  # Round to nearest 10,000 VND
    
    # Generate size quantities
    # Randomly decide if product has size information
    has_size_info = random.choice([True, True, False])  # 2/3 chance of having size info
    
    if has_size_info:
        # Generate random quantities for each size
        size_quantities = {}
        for size in sizes:
            # 1/4 chance of having 0 quantity for a size
            if random.random() < 0.25:
                size_quantities[size] = 0
            else:
                size_quantities[size] = random.randint(1, 50)
        
        # Create size info string
        size_info = ",".join([f"{size}:{qty}" for size, qty in size_quantities.items() if qty > 0])
        
        # Calculate total quantity
        total_quantity = sum(size_quantities.values())
    else:
        size_info = "Không có thông tin"
        total_quantity = random.randint(10, 200)
    
    # Add product to list
    product_data = {
        "ID": "",  # Leave ID blank for import
        "Tên sản phẩm": product_name,
        "Danh mục": category,
        "Giá": price,
        "Tồn kho": total_quantity,
        "Kích thước": size_info
    }
    
    products.append(product_data)

# Create DataFrame
df = pd.DataFrame(products)

# Save to Excel
excel_path = os.path.join(os.path.dirname(__file__), "100_products_for_import.xlsx")
df.to_excel(excel_path, index=False)

# Now create a more formatted version with proper styling
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Format headers
for col_num, column_title in enumerate(df.columns, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add border
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    cell.border = thin_border

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
    adjusted_width = min(max_length + 2, 50)  # Cap width at 50 to avoid too wide columns
    ws.column_dimensions[column].width = adjusted_width

# Save the formatted workbook
formatted_excel_path = os.path.join(os.path.dirname(__file__), "100_products_for_import_formatted.xlsx")
wb.save(formatted_excel_path)

print(f"Excel files created at:\n{excel_path}\n{formatted_excel_path}")
