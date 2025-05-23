import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# Create a new Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Products"

# Define headers
headers = [
    "Tên sản phẩm", 
    "Danh mục", 
    "Giá", 
    "Size S", 
    "Size M", 
    "Size L", 
    "Size XL", 
    "Size 2XL", 
    "Mô tả"
]

# Add headers to the worksheet
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add border
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    cell.border = thin_border

# Add sample data
sample_data = [
    [
        "Veston đen cổ điển wool", 
        "Veston", 
        "2500000", 
        "5", 
        "8", 
        "10", 
        "7", 
        "3", 
        "Veston đen cổ điển wool chất liệu cao cấp, phù hợp cho các dịp trang trọng."
    ],
    [
        "Quần tây xám slim fit", 
        "Quần tây", 
        "850000", 
        "10", 
        "12", 
        "15", 
        "8", 
        "5", 
        "Quần tây xám slim fit, chất liệu mềm mại, thoáng mát, thích hợp đi làm."
    ],
    [
        "Áo sơ mi trắng công sở", 
        "Áo sơ mi", 
        "450000", 
        "15", 
        "20", 
        "18", 
        "12", 
        "8", 
        "Áo sơ mi trắng công sở, chất cotton cao cấp, form dáng vừa vặn."
    ]
]

# Add sample data to the worksheet
for row_num, row_data in enumerate(sample_data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = cell_value
        
        # Add border
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.border = thin_border
        
        # Center align numeric cells
        if col_num in range(3, 9):  # Price and size columns
            cell.alignment = Alignment(horizontal="center")

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save the workbook
file_path = os.path.join(os.path.dirname(__file__), "product_import_template.xlsx")
wb.save(file_path)

print(f"Template created at: {file_path}")

# Now create a file with 100 random products using the same format
import random

# Define categories
categories = ["Veston", "Quần tây", "Áo sơ mi"]

# Define product name templates for each category
veston_templates = [
    "Veston {color} {style} {material}",
    "Áo vest {color} {style} {fit}",
    "Bộ vest {color} {occasion} {material}",
    "Veston {brand} {color} {fit}",
    "Áo blazer {color} {style} {material}"
]

quan_tay_templates = [
    "Quần tây {color} {style} {material}",
    "Quần âu {color} {fit} {material}",
    "Quần tây {brand} {color} {fit}",
    "Quần âu {color} {style} {occasion}",
    "Quần tây {color} {material} {fit}"
]

ao_so_mi_templates = [
    "Áo sơ mi {color} {style} {material}",
    "Áo sơ mi {brand} {color} {fit}",
    "Áo sơ mi {color} {pattern} {occasion}",
    "Áo sơ mi {color} {sleeve} {material}",
    "Áo sơ mi {color} {style} {fit}"
]

# Define attributes for product names
colors = ["đen", "trắng", "xanh navy", "xanh dương", "xám", "be", "nâu", "xanh rêu", "đỏ đô", "xanh đen", "kẻ sọc", "họa tiết"]
styles = ["cổ điển", "hiện đại", "thanh lịch", "trẻ trung", "công sở", "dự tiệc", "casual", "slim fit", "regular fit"]
materials = ["len", "cotton", "linen", "polyester", "wool", "cashmere", "tweed", "kaki", "nhung", "denim"]
brands = ["Elegance", "Gentleman", "Classic", "Modern", "Premium", "Luxury", "Executive", "Business", "Formal"]
fits = ["ôm body", "regular fit", "slim fit", "oversize", "standard fit", "relaxed fit", "skinny fit"]
occasions = ["dự tiệc", "công sở", "cưới hỏi", "dạo phố", "lễ hội", "casual", "formal"]
patterns = ["trơn", "kẻ sọc", "kẻ caro", "họa tiết", "chấm bi", "in hoa", "thêu"]
sleeves = ["tay dài", "tay ngắn", "tay lỡ", "cộc tay"]

# Define price ranges for each category
price_ranges = {
    "Veston": (1500000, 5000000),
    "Quần tây": (500000, 1500000),
    "Áo sơ mi": (350000, 1200000)
}

# Define description templates
description_templates = [
    "{product_name} chất liệu {material} cao cấp, thiết kế {style}, phù hợp cho {occasion}.",
    "{product_name} phong cách {style}, chất {material} mềm mại, thoáng mát, thích hợp {occasion}.",
    "{product_name} thiết kế {fit}, chất liệu {material} bền đẹp, phù hợp cho {occasion}.",
    "{product_name} kiểu dáng {style}, form {fit}, chất {material} cao cấp.",
    "{product_name} thiết kế tinh tế, chất liệu {material}, kiểu dáng {style}, phù hợp {occasion}."
]

# Create a new workbook for 100 products
wb_full = openpyxl.Workbook()
ws_full = wb_full.active
ws_full.title = "Products"

# Add headers
for col_num, header in enumerate(headers, 1):
    cell = ws_full.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# Generate 100 random products
for i in range(2, 102):  # Start from row 2 (after header)
    # Randomly select a category
    category = random.choice(categories)
    
    # Select appropriate template based on category
    if category == "Veston":
        template = random.choice(veston_templates)
    elif category == "Quần tây":
        template = random.choice(quan_tay_templates)
    else:  # Áo sơ mi
        template = random.choice(ao_so_mi_templates)
    
    # Generate product name
    product_name = template.format(
        color=random.choice(colors),
        style=random.choice(styles),
        material=random.choice(materials),
        brand=random.choice(brands),
        fit=random.choice(fits),
        occasion=random.choice(occasions),
        pattern=random.choice(patterns),
        sleeve=random.choice(sleeves)
    )
    
    # Generate price within category range
    min_price, max_price = price_ranges[category]
    price = random.randint(min_price // 10000, max_price // 10000) * 10000  # Round to nearest 10,000 VND
    
    # Generate size quantities
    size_s = random.randint(0, 20)
    size_m = random.randint(0, 20)
    size_l = random.randint(0, 20)
    size_xl = random.randint(0, 20)
    size_2xl = random.randint(0, 20)
    
    # Generate description
    description_template = random.choice(description_templates)
    description = description_template.format(
        product_name=product_name,
        material=random.choice(materials),
        style=random.choice(styles),
        fit=random.choice(fits),
        occasion=random.choice(occasions)
    )
    
    # Add row data
    row_data = [
        product_name,
        category,
        price,
        size_s,
        size_m,
        size_l,
        size_xl,
        size_2xl,
        description
    ]
    
    for col_num, cell_value in enumerate(row_data, 1):
        ws_full.cell(row=i, column=col_num).value = cell_value

# Auto-adjust column widths
for col in ws_full.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
    adjusted_width = min(max_length + 2, 50)  # Cap width at 50 to avoid too wide columns
    ws_full.column_dimensions[column].width = adjusted_width

# Save the workbook with 100 products
file_path_full = os.path.join(os.path.dirname(__file__), "100_products_import_format.xlsx")
wb_full.save(file_path_full)

print(f"100 products file created at: {file_path_full}")
